import oracledb
import openpyxl
import os
import platform
import shutil
import datetime
# import win32com.client as win32
# import pyodbc
import jaydebeapi
from openpyxl.worksheet.table import TableStyleInfo

# --- Helper to get current time string ---
def get_get_current_time():
    return datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")

# --- Get Documents folder cross-platform ---
def get_documents_folder():
    if platform.system() == "Windows":
        print("Running on Windows")
        return os.path.join(os.environ['USERPROFILE'], 'Documents')
    elif platform.system() == "Linux":
        print("Running on Linux")
        return os.path.join(os.path.expanduser("~"), "Documents")
    elif platform.system() == "Darwin":
        print("Running on macOS")
        return os.path.join(os.path.expanduser("~"), "Documents")
    else:
        print(f"Running on an unknown OS: {platform.system()}")

# --- Safe Excel writing to avoid Unicode errors ---
def safe_excel_write(ws, row_idx, col_idx, value):
    if isinstance(value, bytes):
        try:
            value = value.decode("utf-8", errors="replace")
        except:
            value = str(value)
    ws.cell(row=row_idx, column=col_idx, value=value)

# --- Setup paths ---
location = get_documents_folder()
print(location)
DBID_Automation = os.path.join(location, 'DBID_Automation')
dir_path = os.path.join(DBID_Automation, 'Report')
print(dir_path)

# --- Get newest file in dir ---
def newest_file_in_dir(dir_path):
    files = [f for f in os.listdir(dir_path) if os.path.isfile(os.path.join(dir_path, f))]
    files.sort(key=lambda x: os.path.getmtime(os.path.join(dir_path, x)), reverse=True)
    return files[0]

new = newest_file_in_dir(dir_path)
today = datetime.datetime.now().strftime("%m_%d_%y")
template_file = os.path.join(dir_path, new)
new_file = f"{template_file[:-13]}{today}.xlsx"
print(new_file)

# Copy or replace file
if new_file == template_file:
    os.remove(template_file)
    new = newest_file_in_dir(dir_path)
    template_file = os.path.join(dir_path, new)
    shutil.copy2(template_file, new_file)
else:
    shutil.copy2(template_file, new_file)

# --- Oracle Connection ---
try:
    username = "wdbs"
    password = "YBkHzLmvsQG-n3rTUOpCdAfW$"
    dsn = "topoprd"
    tns_admin_dir = "/Users/AF35861/Downloads/instantclient_23_3/network/admin"
    ora_conn = oracledb.connect(user=username, password=password, dsn=dsn, config_dir=tns_admin_dir)
    print("Successfully connected to Oracle Database!")
    ora_cursor = ora_conn.cursor()
    print(f"Oracle Database Version: {ora_conn.version}")
except Exception as err:
    print(f"Error connecting to database: {err}")

finally:
    if 'connection' in locals() and ora_conn:
        ora_conn.close()


# --- SNOW Replica Connection ---
try:
    # jTDS Driver name
    driver_name = "net.sourceforge.jtds.jdbc.Driver"

    # jTDS Connection URL format:
    # jdbc:jtds:<server_type>://<server>[:<port>][/<database>][;<property>=<value>[;...]]
    connection_url = "jdbc:jtds:sqlserver://sql-snow-dev.us.ad.wellpoint.com:10002/ServiceNow;domain=us;ssl=required;socketTimeout=20000;usentlmv2=true;"

    # Connection properties (replace with your credentials)
    connection_properties = {
        "user": "snwihub",
        "password": "NqIHNE72wTI!v9hq-!nNCud9H",
        # Optional: Add other jTDS properties as needed, e.g., "domain": "your_domain"
    }

    # Path to the downloaded jTDS JAR file
    jar_path = "/Users/AF35861/Downloads/jtds-1.3.1.jar" # Replace with your actual path

    # Establish connection
    sql_conn = jaydebeapi.connect(driver_name, connection_url, connection_properties, jar_path)
    print("Successfully connected to SNOW Replica Database!")
    sql_cursor = sql_conn.cursor()

    # Execute the version query
    sql_cursor.execute("SELECT @@VERSION")

    # Fetch and print the result
    version_info = sql_cursor.fetchone()
    if version_info:
        print(f"SQL Server Version: {version_info[0]}")
    else:
        print("Could not retrieve SQL Server version.")

except Exception as err:
    print(f"Error connecting to database: {err}")

finally:
    if 'connection' in locals() and sql_conn:
        sql_conn.close()


# --- Read queries from file ---
Query_Location = os.path.join(DBID_Automation, 'queries.txt')
with open(Query_Location, 'r', encoding='utf-8') as f:
    queries = f.read()
a = queries.split(";")

loop = []

job, error, dbid, snow, tickets = a[0], a[1], a[2], a[3], a[4]
loop.extend([job, error, dbid, snow, tickets])

# --- Loop through all queries ---
for i in loop:
    if i==dbid:
        print(get_get_current_time(),'--     DBID Collection Query Started')
        ora_cursor.execute(dbid)
        # Load the Excel workbook
        workbook = openpyxl.load_workbook(new_file)
        # Get the target sheet
        sheet = workbook['DBID Collection']
        #################
        # Clear data in columns 9 to 28 (I to AB), starting from row 2
        for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row, min_col=9, max_col=27):
            for cell in row:
                cell.value = None

        # Paste the new data into columns 9 to 28 (I to AB), starting from row 2
        for i, row in enumerate(cursor, start=2):
            for j, value in enumerate(row, start=9):
                sheet.cell(row=i, column=j, value=value)

        # # Dynamically adjust the table range to include the entire sheet
        # table = sheet.tables['DBIDS']  # Replace with the actual table name
        # table_end_row = sheet.max_row  # Includes all rows with data
        # table_end_col = sheet.max_column  # Includes all columns (A to AB)

        # # Update the table range to include the entire sheet
        # table.ref = f"A1:{sheet.cell(row=table_end_row, column=table_end_col).coordinate}"

        # # Optionally apply a style to the table
        # style = TableStyleInfo(
        #     name="TableStyleLight2",  # You can choose any style you prefer
        #     showFirstColumn=False,
        #     showLastColumn=False,
        #     showRowStripes=True,
        #     showColumnStripes=False
        # )
        # table.tableStyleInfo = style

        ###############


        workbook.save(new_file)
        cursor.close()
        #connection.close()
        print(get_get_current_time(),'--     DBID Sucessfull')
    elif i==job:
        print(get_get_current_time(),'--     Job Completion Query Started')
        #connection = cx_Oracle.connect("wdbs", "catbarf", "WDBS_PROD")
        # cursor = connection.cursor()
        ora_cursor.execute(job)
        # Load the Excel workbook
        workbook = openpyxl.load_workbook(new_file)
        # Get the target sheet
        sheet = workbook['Job Completion']
        
        # Clear the existing data in the sheet, starting from the second row
        for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row):
            for cell in row:
                cell.value = None

        # Paste the data from the database into the sheet, starting from the second row
        for i, row in enumerate(cursor, start=2):
            for j, cell in enumerate(row, start=1):
                sheet.cell(row=i, column=j, value=cell)

         # Dynamically adjust the table range
        table = sheet.tables['JOB']  # Replace 'YourTableName' with the actual table name in Excel
        table_end_row = sheet.max_row
        table_end_col = sheet.max_column   

        # Update the table range
        table.ref = f"A1:{sheet.cell(row=table_end_row, column=table_end_col).coordinate}"

        # Optionally, apply a style to the table
        style = TableStyleInfo(
            name="TableStyleLight2", 
            showFirstColumn=False,
            showLastColumn=False,
            showRowStripes=True,
            showColumnStripes=False
        )
        table.tableStyleInfo = style

        workbook.save(new_file)
        cursor.close()
        #connection.close()
        print(get_get_current_time(),'--     Job Completion Sucessfull')
    
    elif i==error:
        try:
            print(get_get_current_time(),'--     Error Message Query Started')
            #connection = cx_Oracle.connect("wdbs", "catbarf", "WDBS_PROD")
            # cursor = connection.cursor()
            ora_cursor.execute(error)
            # Load the Excel workbook
            workbook = openpyxl.load_workbook(new_file)
            # Get the target sheet
            sheet = workbook['Error Messages']
            
            # Clear the existing data in the sheet, starting from the second row
            for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row):
                for cell in row:
                    cell.value = None

            # Paste the data from the database into the sheet, starting from the second row
            for i, row in enumerate(ora_cursor, start=2):
                for j, cell in enumerate(row, start=1):
                    sheet.cell(row=i, column=j, value=cell)
            '''
            # Dynamically adjust the table range
            table = sheet.tables['ERRORS']  # Replace 'YourTableName' with the actual table name in Excel
            table_end_row = sheet.max_row
            table_end_col = sheet.max_column   

            # Update the table range
            table.ref = f"A1:{sheet.cell(row=table_end_row, column=table_end_col).coordinate}"

            # Optionally, apply a style to the table
            style = TableStyleInfo(
                name="TableStyleLight2", 
                showFirstColumn=False,
                showLastColumn=False,
                showRowStripes=True,
                showColumnStripes=False
            )
            table.tableStyleInfo = style
            '''

            workbook.save(new_file)
            ora_cursor.close()
            #connection.close()
            print(get_get_current_time(),'--     Error Message Sucessfull')
        except Exception as e:
            print(str(e))
            pass

'''
    elif i==snow:
        print(get_get_current_time(),'--     SNOW Query Started')
        #connection = cx_Oracle.connect("wdbs", "catbarf", "WDBS_PROD")
        cursor_sql = conn.cursor()
        cursor_sql.execute(snow)
        # Load the Excel workbook
        workbook = openpyxl.load_workbook(new_file)
        # Get the target sheet
        sheet = workbook['SNOW Extract']
        
        # Clear the existing data in the sheet, starting from the second row and column 7 to the end
        for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row):
            for cell in row[2:]:
                cell.value = None
                
        # Paste the data from the database into the sheet, starting from the second row and column 7
        for i, row in enumerate(cursor_sql, start=2):
            for j, cell in enumerate(row[0:], start=2):
                sheet.cell(row=i, column=j, value=cell)

        #  # Dynamically adjust the table range
        # table = sheet.tables['SNOW']  # Replace 'YourTableName' with the actual table name in Excel
        # table_end_row = sheet.max_row
        # table_end_col = sheet.max_column   

        # # Update the table range
        # table.ref = f"A1:{sheet.cell(row=table_end_row, column=table_end_col).coordinate}"

        # # Optionally, apply a style to the table
        # style = TableStyleInfo(
        #     name="TableStyleLight2", 
        #     showFirstColumn=False,
        #     showLastColumn=False,
        #     showRowStripes=True,
        #     showColumnStripes=False
        # )
        # table.tableStyleInfo = style
        workbook.save(new_file)
        cursor_sql.close()
        #connection.close()
        print(get_get_current_time(),'--     SNOW Query Successfull')
        '''
'''
    elif i==tickets:
        print(get_get_current_time(),'--     Incident Tickets Query Started')
        #connection = cx_Oracle.connect("wdbs", "catbarf", "WDBS_PROD")
        cursor_sql = conn.cursor()
        cursor_sql.execute(tickets)
        # Load the Excel workbook
        workbook = openpyxl.load_workbook(new_file)
        # Get the target sheet
        sheet = workbook['Tickets']
        
        # Clear the existing data in the sheet, starting from the second row
        for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row):
            for cell in row:
                cell.value = None

        # Paste the data from the database into the sheet, starting from the second row
        for i, row in enumerate(cursor_sql, start=2):
            for j, cell in enumerate(row, start=1):
                sheet.cell(row=i, column=j, value=cell)


        # Dynamically adjust the table range
        table = sheet.tables['TICKETS']  # Replace 'YourTableName' with the actual table name in Excel
        table_end_row = sheet.max_row
        table_end_col = sheet.max_column   

        # Update the table range
        table.ref = f"A1:{sheet.cell(row=table_end_row, column=table_end_col).coordinate}"

        # Optionally, apply a style to the table
        style = TableStyleInfo(
            name="TableStyleLight2", 
            showFirstColumn=False,
            showLastColumn=False,
            showRowStripes=True,
            showColumnStripes=False
        )
        table.tableStyleInfo = style

        workbook.save(new_file)
        cursor_sql.close()
        #connection.close()
        print(get_get_current_time(),'--    Incident Tickets Query Sucessfull')    
        '''

print(get_get_current_time(),"--     Today's file Exported")