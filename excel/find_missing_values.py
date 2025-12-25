import pandas as pd
from openpyxl import load_workbook

def find_missing_values(excel_file, sheet_name, column_a_name, column_b_name):
    """
    Finds values in column A that are not present in column B of an Excel sheet.

    Args:
        excel_file (str): The path to the Excel file.
        sheet_name (str): The name of the sheet to read.
        column_a_name (str): The name of column A.
        column_b_name (str): The name of column B.

    Returns:
        pandas.Series: A Series containing values from column A that are not in column B.
    """
    try:
        df = pd.read_excel(excel_file, sheet_name=sheet_name)
    except FileNotFoundError:
        print(f"Error: The file '{excel_file}' was not found.")
        return pd.Series()
    except Exception as e:
        print(f"Error reading Excel file: {e}")
        return pd.Series()

    if column_a_name not in df.columns or column_b_name not in df.columns:
        print(f"Error: One or both specified columns ('{column_a_name}', '{column_b_name}') not found in the sheet.")
        return pd.Series()

    # Convert columns to sets for efficient comparison
    set_a = set(df[column_a_name].dropna().astype(str))  # Convert to string and drop NaNs
    set_b = set(df[column_b_name].dropna().astype(str))  # Convert to string and drop NaNs

    # Find values in set_a that are not in set_b
    missing_values = list(set_a - set_b)
    # print(f"Found {len(missing_values)} missing values.")

    return pd.Series(missing_values)

# Example usage:
excel_file_path = '/Users/AF35861/Downloads/Teradata_Inventory.xlsx'
# modified_file_path = '/Users/AF35861/Downloads/Modified_Teradata_Inventory.xlsx'
sheet = 'DWTEST2'
col_a = 'DWTEST2'
col_b = 'SERVICENOW'

result = find_missing_values(excel_file_path, sheet, col_a, col_b)
result = result.sort_values().reset_index(drop=True)

if not result.empty:
    result_size = len(result.tolist())
    print(f"Values in '{col_a}' not found in '{col_b}': {result_size}")
else:
    print(f"All values in '{col_a}' are present in '{col_b}', or no data was found.")

sn_result = find_missing_values(excel_file_path, sheet, col_b, col_a)
sn_result = sn_result.sort_values().reset_index(drop=True)

if not sn_result.empty:
    sn_result_size = len(sn_result.tolist())
    print(f"Values in '{col_b}' not found in '{col_a}': {sn_result_size}")
else:
    print(f"All values in '{col_b}' are present in '{col_a}', or no data was found.")

# Not in SN
new_column_data = result.tolist()

workbook = load_workbook(excel_file_path)
worksheet = workbook[sheet]
column_index = 3
worksheet.insert_cols(idx=column_index)

worksheet.cell(row=1, column=3).value = 'Not in SN'

for row_num, item in enumerate(new_column_data, start=2):  # Start at row 2
        worksheet.cell(row=row_num, column=column_index).value = item

workbook.save(excel_file_path)

# Not in Teradata
exs_column_data = sn_result.tolist()

workbook = load_workbook(excel_file_path)
worksheet = workbook[sheet]
column_index = 4
worksheet.insert_cols(idx=column_index)

worksheet.cell(row=1, column=column_index).value = 'Not in Teradata'

for row_num, item in enumerate(exs_column_data, start=2):
        worksheet.cell(row=row_num, column=column_index).value = item

workbook.save(excel_file_path)


